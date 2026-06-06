import io

from django.http import HttpResponse


def build_xlsx_response(filename, headers, rows):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    for idx, header in enumerate(headers, start=1):
        width = max(len(str(header)), 12)
        for row in rows[:200]:
            if idx - 1 < len(row):
                width = max(width, min(len(str(row[idx - 1])), 48))
        ws.column_dimensions[get_column_letter(idx)].width = width + 2
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    if not filename.lower().endswith(".xlsx"):
        filename = f"{filename}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_query(request, fmt="pdf"):
    params = request.GET.copy()
    params["format"] = fmt
    return params.urlencode()
